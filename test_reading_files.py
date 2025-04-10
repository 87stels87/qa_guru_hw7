import codecs
import csv
import zipfile

from openpyxl import load_workbook
from pypdf import PdfReader


def test_reading_pdf(create_archive):
    with zipfile.ZipFile("resources/files.zip", 'r') as archive:
        for info in archive.infolist():
            if info.filename.endswith('.pdf'):
                with archive.open(info.filename) as pdf_file:
                    reader = PdfReader(pdf_file)
                    print(reader.pages[1].extract_text())
    assert "1 MB sample PDF file offers an easy solution for various uses" in reader.pages[1].extract_text()


def test_reading_xlsx(create_archive):
    with zipfile.ZipFile("resources/files.zip", 'r') as archive:
        for info in archive.infolist():
            if info.filename.endswith('.xlsx'):
                with archive.open(info.filename) as xlsx_file:
                    workbook = load_workbook(xlsx_file)
                    sheet = workbook.active
    assert "Austen Russel" in sheet.cell(row=2, column=1).value


def test_reading_csv(create_archive):
    expected_value = [
        ["1;2;3;4;5"],
        ["6;7;8;9;10"],
    ]
    with zipfile.ZipFile("resources/files.zip", 'r') as archive:
        for info in archive.infolist():
            if info.filename.endswith('.csv'):
                with archive.open(info.filename, 'r', ) as csv_file:
                    reader = csv.reader(codecs.iterdecode(csv_file, 'utf-8'))
                    actual_data = list(reader)
    assert actual_data == expected_value